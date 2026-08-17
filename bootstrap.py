#!/usr/bin/env python3
"""bootstrap.py — Bootstrap VentureLab with benchmark research data.

Imports the agent_reality_ventures_benchmark_2026.xlsx research workbook
and sets up the full pipeline for autonomous venture research.
"""
import json
import sys
from pathlib import Path
from datetime import datetime, timezone

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import openpyxl


def load_workbook():
    """Load the benchmark workbook."""
    wb_path = ROOT / "data" / "agent_reality_ventures_benchmark_2026.xlsx"
    return openpyxl.load_workbook(wb_path)


def extract_ventures(wb):
    """Extract ventures from Scorecard sheet."""
    ws = wb["Scorecard"]
    ventures = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            venture = {
                "idea_id": f"VENT_{row[0].replace(' ', '_').replace('/', '_')[:20]}",
                "idea": row[0],
                "scores": {
                    "market_timing": row[1],
                    "pain_severity": row[2],
                    "wtp": row[3],
                    "api_fit": row[4],
                    "whitespace": row[5],
                    "defensibility": row[6],
                    "mvp_buildability": row[7],
                    "expansion": row[8],
                    "standards": row[9],
                    "reg_simplicity": row[10],
                },
                "thesis": row[12],
                "best_wedge": row[13],
                "monetization": row[14],
                "main_risks": row[15],
                "status": "evaluated",
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            ventures.append(venture)
    
    return ventures


def extract_competitors(wb):
    """Extract competitors from Competitors sheet."""
    ws = wb["Competitors"]
    competitors = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            competitor = {
                "venture": row[0],
                "player": row[1],
                "relation": row[2],
                "what_it_does": row[3],
                "capability": row[4],
                "business_model": row[5],
                "strength": row[6],
                "gap": row[7],
                "source": row[8],
            }
            competitors.append(competitor)
    
    return competitors


def extract_oss(wb):
    """Extract OSS projects from OSS_GitHub sheet."""
    ws = wb["OSS_GitHub"]
    oss = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            project = {
                "venture": row[0],
                "repo": row[1],
                "why_relevant": row[2],
                "url": row[3],
            }
            oss.append(project)
    
    return oss


def extract_evidence(wb):
    """Extract evidence from Evidence sheet."""
    ws = wb["Evidence"]
    evidence = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            ev = {
                "theme": row[0],
                "finding": row[1],
                "applies_to": row[2],
                "source": row[3],
            }
            evidence.append(ev)
    
    return evidence


def extract_mvp_roadmap(wb):
    """Extract MVP roadmap from MVP_Roadmap sheet."""
    ws = wb["MVP_Roadmap"]
    roadmap = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            item = {
                "venture": row[0],
                "window": row[1],
                "mvp": row[2],
                "constraint": row[3],
                "monetization": row[4],
            }
            roadmap.append(item)
    
    return roadmap


def extract_criteria(wb):
    """Extract scoring criteria from Benchmark sheet."""
    ws = wb["Benchmark"]
    criteria = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            criterion = {
                "name": row[0],
                "weight": row[1],
                "meaning": row[2],
            }
            criteria.append(criterion)
    
    return criteria


def save_data(data, filename):
    """Save data to JSONL file."""
    filepath = ROOT / "data" / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        for item in data:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")
    return filepath


def main():
    print("=== BOOTSTRAPPING VENTURELAB ===")
    print()
    
    # Load workbook
    print("Loading benchmark workbook...")
    wb = load_workbook()
    print(f"  Sheets: {wb.sheetnames}")
    print()
    
    # Extract data
    print("Extracting ventures...")
    ventures = extract_ventures(wb)
    print(f"  Found {len(ventures)} ventures")
    for v in ventures:
        print(f"    - {v['idea']}")
    print()
    
    print("Extracting competitors...")
    competitors = extract_competitors(wb)
    print(f"  Found {len(competitors)} competitors")
    print()
    
    print("Extracting OSS projects...")
    oss = extract_oss(wb)
    print(f"  Found {len(oss)} OSS projects")
    print()
    
    print("Extracting evidence...")
    evidence = extract_evidence(wb)
    print(f"  Found {len(evidence)} evidence points")
    print()
    
    print("Extracting MVP roadmap...")
    roadmap = extract_mvp_roadmap(wb)
    print(f"  Found {len(roadmap)} roadmap items")
    print()
    
    print("Extracting scoring criteria...")
    criteria = extract_criteria(wb)
    print(f"  Found {len(criteria)} criteria")
    print()
    
    # Save data
    print("Saving data files...")
    save_data(ventures, "ideas.jsonl")
    save_data(competitors, "competitors.jsonl")
    save_data(oss, "oss_projects.jsonl")
    save_data(evidence, "evidence.jsonl")
    save_data(roadmap, "roadmap.jsonl")
    save_data(criteria, "criteria.jsonl")
    print("  All data saved to data/")
    print()
    
    # Create research records
    print("Creating research records...")
    for venture in ventures:
        research_record = {
            "idea_id": venture["idea_id"],
            "idea": venture["idea"],
            "competitors": [c for c in competitors if c["venture"] == venture["idea"]],
            "oss_projects": [o for o in oss if o["venture"] == venture["idea"]],
            "evidence": [e for e in evidence if venture["idea"] in e.get("applies_to", "")],
            "roadmap": [r for r in roadmap if r["venture"] == venture["idea"]],
            "researched_at": datetime.now(timezone.utc).isoformat(),
        }
        
        research_file = ROOT / "data" / "research.jsonl"
        with open(research_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(research_record, ensure_ascii=False) + "\n")
    print(f"  Created research records for {len(ventures)} ventures")
    print()
    
    print("=== BOOTSTRAP COMPLETE ===")
    print()
    print("Next steps:")
    print("  python3 agent/run.py --step report")
    print("  python3 agent/run.py --step watchdog")


if __name__ == "__main__":
    main()
